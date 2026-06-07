"""Export DB tables to Excel (.xlsx)."""

from pathlib import Path
from typing import Optional, Set

from .sav_file import SavFile


def export_to_excel(
    sav: SavFile,
    output_path: Path,
    tables_filter: Optional[Set[str]] = None,
    on_progress=None,
):
    """Export all (or selected) tables to an Excel workbook.

    Args:
        sav: Loaded save file
        output_path: Path for the .xlsx output
        tables_filter: Optional set of short names to export (e.g. {"CZUM"})
        on_progress: Optional callback (table_name, row_count)
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    header_fill = PatternFill(
        start_color="4472C4", end_color="4472C4", fill_type="solid"
    )
    header_font = Font(color="FFFFFF", bold=True)

    table_list = list(sav.db.tables.items())
    if tables_filter:
        table_list = [(k, v) for k, v in table_list if k in tables_filter]

    for short_name, table in table_list:
        if not table.records:
            continue

        sheet_name = (table.long_name or short_name)[:31]
        ws = wb.create_sheet(title=sheet_name)

        headers = [fd.field_name or fd.short_name_str for fd in table.fields]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        total = len(table.records)
        batch_size = 1000
        for batch_start in range(0, total, batch_size):
            batch_end = min(batch_start + batch_size, total)
            for row_offset, record in enumerate(
                table.records[batch_start:batch_end], batch_start
            ):
                row_num = row_offset + 2
                for col_idx, fd in enumerate(table.fields):
                    key = fd.field_name or fd.short_name_str
                    val = record.get(key, "")
                    if isinstance(val, int):
                        pass
                    elif isinstance(val, float):
                        val = round(val, 4)
                    elif val is None:
                        val = ""
                    ws.cell(row=row_num, column=col_idx + 1, value=val)

        if on_progress:
            on_progress(sheet_name, total)

    wb.save(str(output_path))
